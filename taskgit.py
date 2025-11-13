from abc import ABC, abstractmethod
from datetime import datetime
from reportlab.pdfgen import canvas 
from reportlab.lib.pagesizes import letter
import openpyxl

class InvoiceGenerator(ABC):

    def __init__(self, client_name, iteems):
        self.client_name = client_name
        self.items = iteems
        self.date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def calculate_total(self):
        return sum(item['price'] for item in self.items)
    
    @abstractmethod
    def generate_invoice(self):
        pass

class PDFInvoiceGenerator(InvoiceGenerator):

    def generate_invoice(self):
        fileneme =  f"invoce_{self.client_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf = canvas.Canvas(fileneme, pagesize=letter)

        pdf.setFont("Helvetica-Bold", 20)
        pdf.drawString(100, 750, "INVOICE")

        pdf.setFont("Helvetica", 12)
        pdf.drawString(100, 720, f"Client: {self.client_name}")
        pdf.drawString(100, 700, f"Date: {self.date}")

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(100,670,"items:")

        pdf.setFont("Helvetica",11)
        y_position = 650
        for item in self.items:
            pdf.drawString(120, y_position, f"-{item['name']}: ${item['price']}")
            y_position -= 20
        

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(100,y_position - 20, f"Total:${self.calculate_total()}")

        pdf.save()
        return fileneme
    
class ExelInvoiceGenerator(InvoiceGenerator):

    def generate_invoice(self):
        filename = f"invoice_{self.client_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
       
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        ws['A1'] = "INVOICE"
        ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)
        ws['A3'] = f"Client: {self.client_name}"
        ws['A4'] = f"Date: {self.date}"


        ws['A6'] = "Product Name"
        ws['B6'] = "Price"
        ws['A6'].font = openpyxl.styles.Font(bold=True)
        ws['B6'].font = openpyxl.styles.Font(bold=True)

        row = 7
        for item in self.items:
            ws[f'A{row}'] = item['name']
            ws[f'B{row}'] = item['price']
            row += 1


        ws[f'A{row + 1}'] = "TOTAL"
        ws[f'B{row + 1}'] = self.calculate_total()
        ws[f'A{row + 1}'].font = openpyxl.styles.Font(bold= True)
        ws[f'B{row + 1}'].font = openpyxl.styles.Font(bold= True)


        ws[f'A{row + 3}'] = f"Created: {self.date}"
        wb.save(filename)
        return filename
    
class HTMLInvoiceGenerator(InvoiceGenerator):
    def generate_invoice(self):
        filename = f"invoice_{self.client_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.html"
        
        
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Invoice - {self.client_name}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #333; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
        .total {{ font-weight: bold; font-size: 18px; }}
    </style>
</head>
<body>
    <h1>INVOICE</h1>
    <p><strong>Client:</strong> {self.client_name}</p>
    <p><strong>Date:</strong> {self.date}</p>
    
    <table>
        <tr>
            <th>Product Name</th>
            <th>Price</th>
        </tr>
"""
        
       
        for item in self.items:
            html_content += f"""
        <tr>
            <td>{item['name']}</td>
            <td>${item['price']}</td>
        </tr>
"""
        
      
        html_content += f"""
        <tr class="total">
            <td>TOTAL:</td>
            <td>${self.calculate_total()}</td>
        </tr>
    </table>
</body>
</html>
"""
        
       
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return filename
    
class InvoiceManager:
    def __init__(self, generator):
        self.generator = generator

    def create_invoice(self):
        filename = self.generator.generate_invoice()
        print(f"Invoice successfully generated: {filename}")
        return filename
    
if __name__ == "__main__":
    print("="*60)
    print("FINTRACK CO. -INVOICE GENERATOR")
    print("="*60)

    client = "Siddiq Karimov"
    items = [
        {"name": "Laptop", "price": 1000},
        {"name": "Mouse", "price": 50},
        {"name": "Keyboard", "price": 80}
    ]

    print(f"\nClient: {client}")
    print(f"Items: {len(items)} products")
    print(f"Total: ${sum(item['price'] for item in items)}")
    print("\nGenerating invoices...\n")


    pdf_gen = PDFInvoiceGenerator(client, items)
    pdf_manager = InvoiceManager(pdf_gen)
    pdf_manager.create_invoice()


    excel_gen = ExelInvoiceGenerator(client, items)
    excel_manager = InvoiceManager(excel_gen)
    excel_manager.create_invoice()


    html_gen = HTMLInvoiceGenerator(client, items)
    html_manager = InvoiceManager(html_gen)
    html_manager.create_invoice()


    print("\n" + "="*60)
    print("All invoices generated successfully")
    print("="*60)